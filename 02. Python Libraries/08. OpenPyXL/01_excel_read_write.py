# 01 — openpyxl (Excel .xlsx)
# Run: python 01_excel_read_write.py
# Install: pip install openpyxl

from openpyxl import Workbook, load_workbook
import os

path = "students.xlsx"

# --- 1. Write workbook ---
wb = Workbook()
ws = wb.active
ws.title = "Scores"
ws.append(["Name", "Math", "Science"])
ws.append(["Asha", 90, 88])
ws.append(["Ravi", 76, 82])
ws.append(["Meera", 92, 95])
wb.save(path)
print("wrote", path)

# --- 2. Read workbook ---
wb2 = load_workbook(path)
sheet = wb2["Scores"]
print("rows:")
for row in sheet.iter_rows(min_row=2, values_only=True):
    print(row)

# --- 3. Cell access ---
print("B2:", sheet["B2"].value)

wb2.close()
os.remove(path)
