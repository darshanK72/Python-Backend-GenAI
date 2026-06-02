# 02 — Multiple sheets and cell ranges
# Run: python 02_formulas_and_sheets.py

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

wb = Workbook()
summary = wb.active
summary.title = "Summary"
summary["A1"] = "Total Sales"
summary["B1"] = 12500

detail = wb.create_sheet("Detail")
detail.append(["Product", "Qty", "Price"])
detail.append(["Pen", 100, 20])
detail.append(["Book", 50, 150])

# Column width
detail.column_dimensions[get_column_letter(1)].width = 15

wb.save("report.xlsx")
print("sheets:", wb.sheetnames)

from openpyxl import load_workbook
wb = load_workbook("report.xlsx")
print("Summary B1:", wb["Summary"]["B1"].value)
wb.close()
os.remove("report.xlsx")
